from django.utils.translation import gettext_lazy as _
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelRenderer:
    """Renderer Excel (XLSX) pour l'export comptable."""

    HEADERS = [
        "Date Paiement",
        "ID Commande",
        "Moyen Paiement",
        "Montant Brut",
        "TVA Collectée",
        str(_("Total PSP Fees")),
        "Détail Frais",
        "Montant Net",
        "Devise",
        "ID Transaction PSP",
        "Settlement ID",
        "Statut",
    ]

    def render(self, export_data, totals, form_data):
        """
        Génère un fichier Excel.

        Args:
            export_data: Liste de dicts avec les données
            totals: Dict des totaux
            form_data: Paramètres d'export

        Returns:
            bytes: Contenu Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Comptable"

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True)

        # En-têtes
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Données
        for row_num, row_data in enumerate(export_data, 2):
            ws.cell(
                row=row_num, column=1, value=row_data["date_paiement"].strftime("%Y-%m-%d %H:%M:%S")
            )
            ws.cell(row=row_num, column=2, value=row_data["id_commande"])
            ws.cell(row=row_num, column=3, value=row_data["moyen_paiement"])
            ws.cell(row=row_num, column=4, value=float(row_data["montant_brut"]))
            ws.cell(row=row_num, column=5, value=float(row_data["tva_collectee"]))
            ws.cell(row=row_num, column=6, value=float(row_data["frais_psp_total"]))
            ws.cell(row=row_num, column=7, value=row_data["detail_frais"])
            ws.cell(row=row_num, column=8, value=float(row_data["montant_net"]))
            ws.cell(row=row_num, column=9, value=row_data["devise"])
            ws.cell(row=row_num, column=10, value=row_data["id_transaction_psp"])
            ws.cell(row=row_num, column=11, value=row_data["settlement_id"])
            ws.cell(row=row_num, column=12, value=row_data["statut"])

        # Ligne vide
        current_row = len(export_data) + 2

        # Section Totaux Globaux
        current_row += 1
        ws.cell(row=current_row, column=1, value="=== TOTAUX GLOBAUX ===")
        ws.cell(row=current_row, column=1).font = total_font
        ws.cell(row=current_row, column=1).fill = total_fill

        current_row += 1
        ws.cell(row=current_row, column=1, value="Nombre de paiements")
        ws.cell(row=current_row, column=2, value=totals["global"]["count"])

        current_row += 1
        ws.cell(row=current_row, column=1, value="Montant Brut Total")
        ws.cell(row=current_row, column=2, value=float(totals["global"]["montant_brut"]))

        current_row += 1
        ws.cell(row=current_row, column=1, value=str(_("Total VAT Collected")))
        ws.cell(row=current_row, column=2, value=float(totals["global"]["tva_collectee"]))

        current_row += 1
        ws.cell(row=current_row, column=1, value=str(_("Total PSP Fees")))
        ws.cell(row=current_row, column=2, value=float(totals["global"]["frais_psp_total"]))

        current_row += 1
        ws.cell(row=current_row, column=1, value="Montant Net Total")
        ws.cell(row=current_row, column=2, value=float(totals["global"]["montant_net"]))

        # Section Totaux par PSP
        current_row += 2
        ws.cell(row=current_row, column=1, value="=== TOTAUX PAR MOYEN DE PAIEMENT ===")
        ws.cell(row=current_row, column=1).font = total_font
        ws.cell(row=current_row, column=1).fill = total_fill

        for provider, provider_totals in totals["by_provider"].items():
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"--- {provider} ---")
            ws.cell(row=current_row, column=1).font = Font(bold=True)

            current_row += 1
            ws.cell(row=current_row, column=1, value="Nombre")
            ws.cell(row=current_row, column=2, value=provider_totals["count"])

            current_row += 1
            ws.cell(row=current_row, column=1, value="Montant Brut")
            ws.cell(row=current_row, column=2, value=float(provider_totals["montant_brut"]))

            current_row += 1
            ws.cell(row=current_row, column=1, value=str(_("VAT Collected")))
            ws.cell(row=current_row, column=2, value=float(provider_totals["tva_collectee"]))

            current_row += 1
            ws.cell(row=current_row, column=1, value=str(_("PSP Fees")))
            ws.cell(row=current_row, column=2, value=float(provider_totals["frais_psp_total"]))

            current_row += 1
            ws.cell(row=current_row, column=1, value="Montant Net")
            ws.cell(row=current_row, column=2, value=float(provider_totals["montant_net"]))

        # Section Contrôle
        current_row += 2
        ws.cell(row=current_row, column=1, value="=== CONTRÔLE COMPTABLE ===")
        ws.cell(row=current_row, column=1).font = total_font
        ws.cell(row=current_row, column=1).fill = total_fill

        controle_brut_moins_tva = (
            totals["global"]["montant_brut"] - totals["global"]["tva_collectee"]
        )
        controle_final = controle_brut_moins_tva - totals["global"]["frais_psp_total"]

        current_row += 1
        ws.cell(row=current_row, column=1, value="Brut - TVA")
        ws.cell(row=current_row, column=2, value=float(controle_brut_moins_tva))

        current_row += 1
        ws.cell(row=current_row, column=1, value="(Brut - TVA) - Frais")
        ws.cell(row=current_row, column=2, value=float(controle_final))

        current_row += 1
        ws.cell(row=current_row, column=1, value="Net attendu")
        ws.cell(row=current_row, column=2, value=float(totals["global"]["montant_net"]))

        current_row += 1
        ws.cell(row=current_row, column=1, value=str(_("Difference")))
        ws.cell(
            row=current_row,
            column=2,
            value=float(abs(controle_final - totals["global"]["montant_net"])),
        )

        # Ajuster largeurs colonnes
        for col_num in range(1, len(self.HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        # Sauvegarder dans un buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
